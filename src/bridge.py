import asyncio
import json
import os
import uuid
import base64
import tempfile
from pathlib import Path
from typing import Dict, Any
from fastapi import WebSocket

# Using the server_log defined in src/logger.py
from src.logger import server_log as log

BRIDGE_TIMEOUT = float(os.getenv("BRIDGE_TIMEOUT", "60.0"))
LOOPBACK_TIMEOUT = float(os.getenv("LOOPBACK_TIMEOUT", "3600.0"))

class OutlookBridgeManager:
    def __init__(self):
        # Map: user_id -> active WebSocket
        self.active_connections: Dict[str, WebSocket] = {}

        # Map: request_id -> Future (the "semaphore" for waiting)
        self.pending_requests: Dict[str, asyncio.Future] = {}

        # Map: user_id -> client_type (set during client/identify handshake)
        self.client_types: Dict[str, str] = {}

    async def connect(self, websocket: WebSocket, user_id: str):
        await websocket.accept()
        self.active_connections[user_id] = websocket
        log.info(f"🔌 Bridge: User '{user_id}' connected via WebSocket.")

    def disconnect(self, user_id: str):
        if user_id in self.active_connections:
            del self.active_connections[user_id]
        self.client_types.pop(user_id, None)

        # Fail-fast: cancel all pending requests for this user
        failed = []
        for req_id, future in list(self.pending_requests.items()):
            if not future.done():
                future.set_result(f"Outlook connection lost for '{user_id}'.")
                failed.append(req_id)
        for req_id in failed:
            del self.pending_requests[req_id]

        log.info(f"🔌 Bridge: User '{user_id}' disconnected. {len(failed)} pending requests cancelled.")

    async def send_mcp_request(self, user_id: str, tool_name: str, args: dict) -> Any:
        """
        Sends a command to Outlook and waits for the response (blocking execution here).
        """
        if user_id not in self.active_connections:
            msg = f"⚠️ Outlook usage attempt failed: No client connected for user '{user_id}'."
            log.warning(msg)
            return msg

        # 1. Create a unique ID for the request
        request_id = str(uuid.uuid4())

        # 2. Prepare the JSON-RPC packet (MCP Standard)
        payload = {
            "jsonrpc": "2.0",
            "method": "tools/call",
            "params": {
                "name": tool_name,
                "arguments": args
            },
            "id": request_id
        }

        # 3. Create the response "Promise"
        loop = asyncio.get_running_loop()
        future = loop.create_future()
        self.pending_requests[request_id] = future

        try:
            # 4. Send to WebSocket
            ws = self.active_connections[user_id]

            # --- TX LOGGING ---
            json_str = json.dumps(payload)
            await ws.send_text(json_str)
            # Log at INFO level to see it in server.log with full JSON
            log.info(f"📤 MCP TX [{user_id}]: {json_str}")
            # ---------------------------

            # 5. Wait for the response
            result = await asyncio.wait_for(future, timeout=BRIDGE_TIMEOUT)
            return result

        except asyncio.TimeoutError:
            if request_id in self.pending_requests:
                del self.pending_requests[request_id]
            err_msg = f"⚠️ Bridge Timeout: Outlook for {user_id} did not respond within {BRIDGE_TIMEOUT}s."
            log.error(err_msg)
            return err_msg

        except Exception as e:
            err_msg = f"⚠️ Critical Bridge Error: {str(e)}"
            log.error(err_msg, exc_info=True)
            return err_msg

    async def handle_incoming_message(self, user_id: str, message: str):
        """Receives responses from Outlook and unblocks pending requests."""

        try:
            data = json.loads(message)

            # Skip verbose logging for heartbeats
            if data.get("method") == "heartbeat":
                log.debug(f"Heartbeat from {user_id}")
                return

            # RX logging (truncate large messages with attachments)
            log_preview = message[:500] + "..." if len(message) > 500 else message
            log.info(f"MCP RX [{user_id}]: {log_preview}")

            # Client identification handshake — client declares its type, server pushes config
            if data.get("method") == "client/identify":
                request_id = data.get("id", str(uuid.uuid4()))
                client_type = data.get("params", {}).get("client_type", "unknown")
                client_version = data.get("params", {}).get("client_version", "?")
                self.client_types[user_id] = client_type
                config_block = self._build_client_config(client_type)
                response = {
                    "jsonrpc": "2.0",
                    "id": request_id,
                    "result": {"status": "ok", "config": config_block}
                }
                ws = self.active_connections.get(user_id)
                if ws:
                    await ws.send_text(json.dumps(response))
                if config_block:
                    log.info(f"Bridge: '{user_id}' identified as '{client_type}' v{client_version} — config pushed ({len(config_block)} keys)")
                else:
                    log.warning(f"Bridge: '{user_id}' identified as unknown client_type '{client_type}' — no config pushed")
                return

            # Virtual Loopback: email sent to ArcumAI by the plugin
            if data.get("method") == "virtual_loopback/send_email":
                request_id = data.get("id", str(uuid.uuid4()))
                params = data.get("params", {})
                log.info(f"VirtualLoopback: Email received from {user_id} | Subject: '{params.get('subject', '?')}'")

                # Send immediate acknowledgment
                ack = {
                    "jsonrpc": "2.0",
                    "id": request_id,
                    "result": {"status": "processing", "message": "Email received, processing..."}
                }
                ws = self.active_connections.get(user_id)
                if ws:
                    await ws.send_text(json.dumps(ack))

                # Process asynchronously (don't block the WebSocket loop)
                asyncio.create_task(
                    self._process_loopback_email(user_id, request_id, params)
                )
                return

            # Response to one of our tool calls (has an ID matching a pending request)
            if "id" in data:
                req_id = data["id"]
                if req_id in self.pending_requests:
                    future = self.pending_requests[req_id]

                    if "error" in data:
                        log.warning(f"Outlook Error (req {req_id}): {data['error']}")
                        future.set_result(f"Error from Outlook: {data['error']}")
                    else:
                        future.set_result(data.get("result", "OK"))

                    del self.pending_requests[req_id]
                    return

            # Push Events (Notifications from client)
            if "method" in data:
                method = data["method"]
                if method == "closing":
                    log.info(f"Client {user_id} closing")
                else:
                    log.info(f"Notification from Outlook ({user_id}): {method}")

        except Exception as e:
            log.error(f"Error parsing message from {user_id}: {e}", exc_info=True)

    # ------------------------------------------------------------------
    #  VIRTUAL LOOPBACK: PROCESS EMAIL FROM PLUGIN
    # ------------------------------------------------------------------

    def _build_client_config(self, client_type: str) -> dict:
        """Return the config block to push for the given client_type.
        Returns an empty dict for unknown types (client uses its own defaults)."""
        if client_type == "vsto_outlook":
            from src.config import (
                VSTO_MAX_ATTACHMENT_MB, VSTO_MAX_TOTAL_MB,
                VSTO_ARCUMAI_EMAIL, VSTO_ARCUMAI_DISPLAY_NAME,
                VSTO_LOOPBACK_TIMEOUT_MS, VSTO_ENABLE_VIRTUAL_LOOPBACK,
                VSTO_SHOW_NOTIFICATION,
            )
            return {
                "max_attachment_size_mb":     VSTO_MAX_ATTACHMENT_MB,
                "max_total_attachments_mb":   VSTO_MAX_TOTAL_MB,
                "arcumai_email":              VSTO_ARCUMAI_EMAIL,
                "arcumai_display_name":       VSTO_ARCUMAI_DISPLAY_NAME,
                "loopback_timeout_ms":        VSTO_LOOPBACK_TIMEOUT_MS,
                "enable_virtual_loopback":    VSTO_ENABLE_VIRTUAL_LOOPBACK,
                "show_processing_notification": VSTO_SHOW_NOTIFICATION,
            }
        # Future client types: add elif blocks here
        return {}

    async def _process_loopback_email(self, user_id: str, request_id: str, params: dict):
        """
        Process a virtual loopback email:
        1. Extract text content (Phase 2 will add attachment decoding)
        2. Route to the appropriate AI engine
        3. Send response back to the plugin
        """
        subject = params.get("subject", "")
        body = params.get("body", "")
        conversation_id = params.get("conversation_id", "")
        has_attachments = params.get("has_attachments", False)
        cc_recipients = params.get("cc_recipients", [])
        attachments = params.get("attachments", [])
        skipped_attachments = params.get("skipped_attachments", [])

        # Use server-side config constants — single source of truth (pushed to client at connect)
        from src.config import VSTO_MAX_ATTACHMENT_MB as max_attachment_size_mb, \
                               VSTO_MAX_TOTAL_MB as max_total_attachments_mb

        try:
            log.info(f"VirtualLoopback [{user_id}]: Processing | Subject='{subject}' | "
                     f"Attachments={len(attachments)} | Skipped={len(skipped_attachments)} | CC={len(cc_recipients)}")

            # Extract text from each attachment
            attachment_texts = []
            for att in attachments:
                try:
                    att_text = self._process_attachment(att)
                    if att_text:
                        fname = att.get("file_name", "unknown")
                        attachment_texts.append(f"--- FILE: {fname} ---\n{att_text}\n--- END FILE ---")
                except Exception as e:
                    fname = att.get("file_name", "?")
                    log.error(f"VirtualLoopback [{user_id}]: Error processing attachment {fname}: {e}")
                    attachment_texts.append(f"--- FILE: {fname} ---\n[Read error: {e}]\n--- END ---")

            # Build attachment context (document text only, NOT the email body)
            attachment_context = "\n\n".join(attachment_texts) if attachment_texts else ""

            # Append a note about files that were too large to send
            if skipped_attachments:
                skipped_note = (
                    "[NOTE: The following attachments exceeded the size limit and could not be processed: "
                    + ", ".join(skipped_attachments) + "]"
                )
                attachment_context = (attachment_context + "\n\n" + skipped_note).strip() if attachment_context else skipped_note
                log.warning(f"VirtualLoopback [{user_id}]: {len(skipped_attachments)} attachment(s) were skipped: {skipped_attachments}")

            # If the user attached files but ALL were skipped, return a clear error immediately.
            # Do not fall back to RAG silently, as the user clearly wanted document analysis.
            if has_attachments and not attachment_texts and skipped_attachments:
                skipped_list = "\n".join(f"  • {s}" for s in skipped_attachments)
                response_text = (
                    f"Your email could not be processed because all attachments exceeded the configured size limits "
                    f"(max {max_attachment_size_mb} MB per file, {max_total_attachments_mb} MB total).\n\n"
                    f"Files that were too large:\n{skipped_list}\n\n"
                    f"Please compress the files or split them into smaller parts and try again."
                )
                response_html = self._markdown_to_html(response_text)
                response_payload = {
                    "jsonrpc": "2.0",
                    "method": "virtual_loopback/response",
                    "params": {
                        "request_id": request_id,
                        "subject": subject,
                        "conversation_id": conversation_id,
                        "response_text": response_text,
                        "response_html": response_html,
                    }
                }
                ws = self.active_connections.get(user_id)
                if ws:
                    await ws.send_text(json.dumps(response_payload))
                    log.info(f"VirtualLoopback [{user_id}]: Sent size-limit error response for '{subject}'")
                return

            # Route to AI engine (body and attachment_context are kept separate)
            has_real_attachments = has_attachments and len(attachments) > 0
            response_text = await self._route_to_ai_engine(
                user_id, subject, body, attachment_context, use_rag=not has_real_attachments
            )

            # Add CC disclaimer if needed
            if cc_recipients:
                response_text = self._build_cc_disclaimer(cc_recipients) + "\n\n" + response_text

            # Convert markdown to HTML
            response_html = self._markdown_to_html(response_text)

            # Send response back to plugin
            response_payload = {
                "jsonrpc": "2.0",
                "method": "virtual_loopback/response",
                "params": {
                    "request_id": request_id,
                    "subject": subject,
                    "conversation_id": conversation_id,
                    "response_text": response_text,
                    "response_html": response_html,
                }
            }

            ws = self.active_connections.get(user_id)
            if ws:
                await ws.send_text(json.dumps(response_payload))
                log.info(f"VirtualLoopback [{user_id}]: Response sent for '{subject}'")
            else:
                log.error(f"VirtualLoopback [{user_id}]: Connection lost, cannot send response")

        except Exception as e:
            log.error(f"VirtualLoopback [{user_id}]: Processing failed: {e}", exc_info=True)

            # Send error response
            error_payload = {
                "jsonrpc": "2.0",
                "method": "virtual_loopback/response",
                "params": {
                    "request_id": request_id,
                    "subject": subject,
                    "conversation_id": conversation_id,
                    "response_text": f"An error occurred during processing: {str(e)}",
                    "response_html": f"<p style='color:red'>Error: {str(e)}</p>",
                }
            }
            ws = self.active_connections.get(user_id)
            if ws:
                try:
                    await ws.send_text(json.dumps(error_payload))
                except Exception:
                    log.error(f"VirtualLoopback [{user_id}]: Failed to send error response")

    def _process_attachment(self, att_data: dict) -> str:
        """Decode base64 attachment and extract text content."""
        file_name = att_data.get("file_name", "unknown")
        content_b64 = att_data.get("content_base64", "")

        if not content_b64:
            return f"[Empty attachment: {file_name}]"

        file_bytes = base64.b64decode(content_b64)
        ext = Path(file_name).suffix.lower()

        # Save to temporary file for processing
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False, prefix="arcumai_lb_") as tmp:
            tmp.write(file_bytes)
            tmp_path = Path(tmp.name)

        try:
            if ext == ".pdf":
                from src.readers import SmartPDFReader
                reader = SmartPDFReader()
                docs = reader.load_data(tmp_path)
                return "\n".join([d.text for d in docs])

            elif ext == ".msg":
                from src.readers import MyOutlookReader
                reader = MyOutlookReader()
                docs = reader.load_data(tmp_path)
                return "\n".join([d.text for d in docs])

            elif ext == ".eml":
                from src.readers import MyEmlReader
                reader = MyEmlReader()
                docs = reader.load_data(tmp_path)
                return "\n".join([d.text for d in docs])

            elif ext in (".txt", ".csv", ".md"):
                return file_bytes.decode("utf-8", errors="ignore")

            elif ext == ".docx":
                try:
                    import docx
                    doc = docx.Document(tmp_path)
                    return "\n".join([p.text for p in doc.paragraphs])
                except ImportError:
                    return f"[DOCX support requires python-docx: {file_name}]"

            elif ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(tmp_path, data_only=True)
                    text_parts = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        text_parts.append(f"[Sheet: {sheet_name}]")
                        for row in ws.iter_rows(values_only=True):
                            row_text = " | ".join([str(c) if c is not None else "" for c in row])
                            if row_text.strip():
                                text_parts.append(row_text)
                    return "\n".join(text_parts)
                except ImportError:
                    return f"[Excel support requires openpyxl: {file_name}]"

            else:
                return f"[Unsupported file type: {ext}]"

        finally:
            try:
                tmp_path.unlink()
            except Exception:
                pass

    async def _route_to_ai_engine(
        self,
        user_id: str,
        subject: str,
        body: str,
        attachment_context: str = "",
        use_rag: bool = True
    ) -> str:
        """Route the loopback request to the appropriate AI engine.

        For RAG (no attachments): optimize email body as a search query, search knowledge base.
        For FILE_READER (with attachments): use body as the question/instruction,
            attachment_context as the document content to analyse.
        """
        from src.auth import load_users
        from src.engine import UserSession, optimize_prompt_for_rag

        # Find the ArcumAI username from the outlook_id
        users = load_users()
        username = None
        role = "DEFAULT"
        for uname, udata in users.items():
            if udata.get("outlook_id") == user_id:
                username = uname
                role = udata.get("role", "DEFAULT")
                break

        if not username:
            return f"Error: No ArcumAI user associated with outlook_id '{user_id}'."

        # Create a temporary session for this request
        session = UserSession(username=username, role=role)

        if use_rag:
            # No attachments: optimise the email body as a RAG search query
            log.info(f"VirtualLoopback [{user_id}]: Incoming email | Subject='{subject}' | Body ({len(body)} chars):\n{body}")

            try:
                optimized_query = await optimize_prompt_for_rag(subject, body)
                log.info(f"VirtualLoopback [{user_id}]: Optimized prompt:\n{optimized_query}")
            except Exception as e:
                log.warning(f"VirtualLoopback [{user_id}]: Prompt optimization failed ({e}), using raw email as fallback")
                optimized_query = f"Email Subject: {subject}\n\n{body}"

            query = f"@rag {optimized_query}"
            mode_override = None
            log.info(f"VirtualLoopback [{user_id}]: Routing to RAG engine (no attachments)")
        else:
            # Has attachments: body = question, attachment_context = document text
            # The body goes through optimize_prompt_for_rag to strip email noise and extract the
            # core instruction (NER masking also applied if configured).
            log.info(f"VirtualLoopback [{user_id}]: Incoming email with attachments | Subject='{subject}' | "
                     f"Body ({len(body)} chars) | Attachment context ({len(attachment_context)} chars)")
            try:
                optimized_instruction = await optimize_prompt_for_rag(subject, body)
                log.info(f"VirtualLoopback [{user_id}]: Optimized instruction:\n{optimized_instruction}")
            except Exception as e:
                log.warning(f"VirtualLoopback [{user_id}]: Prompt optimization failed ({e}), using raw body as fallback")
                optimized_instruction = f"Subject: {subject}\n\n{body}"

            query = optimized_instruction          # What the user wants done with the document(s)
            session.uploaded_context = attachment_context  # Document text only (no body duplication)
            mode_override = "FILE_READER"
            log.info(f"VirtualLoopback [{user_id}]: Routing to FILE_READER engine "
                     f"(instruction: {len(query)} chars | context: {len(attachment_context)} chars)")

        try:
            response_obj, response_text, used_mode = await session.run_chat_action(query, mode_override=mode_override)
            log.info(f"VirtualLoopback [{user_id}]: AI responded via {used_mode} mode")
            return response_text
        except Exception as e:
            log.error(f"VirtualLoopback [{user_id}]: AI engine error: {e}", exc_info=True)
            return f"Error during AI processing: {str(e)}"

    def _build_cc_disclaimer(self, cc_recipients: list) -> str:
        """Build a disclaimer for emails where ArcumAI was in CC with real recipients."""
        names = ", ".join(cc_recipients)
        return (
            "------------------------------------------------------------\n"
            "NOTE: This response is for you only.\n"
            f"The CC recipients of the original email ({names}) "
            "did NOT receive this analysis.\n"
            "If you find it useful, you can forward this email to them.\n"
            "------------------------------------------------------------"
        )

    def _markdown_to_html(self, markdown_text: str) -> str:
        """Convert markdown text to HTML for email display."""
        try:
            import markdown
            return markdown.markdown(
                markdown_text,
                extensions=['tables', 'fenced_code', 'nl2br']
            )
        except ImportError:
            # Fallback: basic conversion
            import html as html_module
            escaped = html_module.escape(markdown_text)
            return escaped.replace('\n', '<br>')


# Global instance
bridge_manager = OutlookBridgeManager()
