# Copyright (c) 2026 Nicolas Brianza
# Licensed under the MIT License. See LICENSE file in the project root.
from __future__ import annotations

import asyncio
import base64
import json
import tempfile
from pathlib import Path

from src.logger import server_log as log
from src.bridge.pending_results import PendingResultStore


class LoopbackProcessor:
    """
    Handles loopback email processing: attachment decoding, AI routing,
    and response dispatch. Extracted from OutlookBridgeManager.
    Holds a live reference to the manager's active_connections dict.
    """

    def __init__(self, active_connections: dict, pending: PendingResultStore):
        self.active_connections = active_connections  # live reference to manager dict
        self._pending = pending

    async def _process_loopback_email(self, user_id: str, request_id: str, params: dict):
        """
        Process a virtual loopback email:
        1. Extract text content (Phase 2 will add attachment decoding)
        2. Route to the appropriate AI engine
        3. Send response back to the plugin
        """
        subject = params.get("subject", "")
        body = params.get("body", "")
        conversation_id = params.get("conversation_id", "")
        original_message_id = params.get("original_message_id", "")
        has_attachments = params.get("has_attachments", False)
        cc_recipients = params.get("cc_recipients", [])
        attachments = params.get("attachments", [])
        skipped_attachments = params.get("skipped_attachments", [])

        # Use server-side config constants — single source of truth (pushed to client at connect)
        from src.config import VSTO_MAX_ATTACHMENT_MB as max_attachment_size_mb, \
                               VSTO_MAX_TOTAL_MB as max_total_attachments_mb

        try:
            log.info(f"VirtualLoopback [{user_id}]: Processing | Subject='{subject}' | "
                     f"Attachments={len(attachments)} | Skipped={len(skipped_attachments)} | CC={len(cc_recipients)}")

            # Extract text from each attachment
            attachment_texts = []
            for att in attachments:
                try:
                    att_text = self._process_attachment(att)
                    if att_text:
                        fname = att.get("file_name", "unknown")
                        attachment_texts.append(f"--- FILE: {fname} ---\n{att_text}\n--- END FILE ---")
                except Exception as e:
                    fname = att.get("file_name", "?")
                    log.error(f"VirtualLoopback [{user_id}]: Error processing attachment {fname}: {e}")
                    attachment_texts.append(f"--- FILE: {fname} ---\n[Read error: {e}]\n--- END ---")

            # Build attachment context (document text only, NOT the email body)
            attachment_context = "\n\n".join(attachment_texts) if attachment_texts else ""

            # Append a note about files that were too large to send
            if skipped_attachments:
                skipped_note = (
                    "[NOTE: The following attachments exceeded the size limit and could not be processed: "
                    + ", ".join(skipped_attachments) + "]"
                )
                attachment_context = (attachment_context + "\n\n" + skipped_note).strip() if attachment_context else skipped_note
                log.warning(f"VirtualLoopback [{user_id}]: {len(skipped_attachments)} attachment(s) were skipped: {skipped_attachments}")

            # If the user attached files but ALL were skipped, return a clear error immediately.
            # Do not fall back to RAG silently, as the user clearly wanted document analysis.
            if has_attachments and not attachment_texts and skipped_attachments:
                skipped_list = "\n".join(f"  • {s}" for s in skipped_attachments)
                response_text = (
                    f"Your email could not be processed because all attachments exceeded the configured size limits "
                    f"(max {max_attachment_size_mb} MB per file, {max_total_attachments_mb} MB total).\n\n"
                    f"Files that were too large:\n{skipped_list}\n\n"
                    f"Please compress the files or split them into smaller parts and try again."
                )
                response_html = self._markdown_to_html(response_text)
                response_params = {
                    "request_id": request_id,
                    "subject": subject,
                    "conversation_id": conversation_id,
                    "original_message_id": original_message_id,
                    "response_text": response_text,
                    "response_html": response_html,
                }
                ws = self.active_connections.get(user_id)
                if ws:
                    await ws.send_text(json.dumps({"jsonrpc": "2.0",
                                                   "method": "virtual_loopback/response",
                                                   "params": response_params}))
                    log.info(f"VirtualLoopback [{user_id}]: Sent size-limit error response for '{subject}'")
                else:
                    await self._pending.save(user_id, request_id, conversation_id, response_params)
                return

            # Route to AI engine (body and attachment_context are kept separate)
            has_real_attachments = has_attachments and len(attachments) > 0
            response_text = await self._route_to_ai_engine(
                user_id, subject, body, attachment_context, use_rag=not has_real_attachments
            )

            # Add CC disclaimer if needed
            if cc_recipients:
                response_text = self._build_cc_disclaimer(cc_recipients) + "\n\n" + response_text

            # Convert markdown to HTML
            response_html = self._markdown_to_html(response_text)

            # Send response back to plugin (or store if client is offline)
            response_params = {
                "request_id": request_id,
                "subject": subject,
                "conversation_id": conversation_id,
                "original_message_id": original_message_id,
                "response_text": response_text,
                "response_html": response_html,
            }
            ws = self.active_connections.get(user_id)
            if ws:
                await ws.send_text(json.dumps({"jsonrpc": "2.0",
                                               "method": "virtual_loopback/response",
                                               "params": response_params}))
                log.info(f"VirtualLoopback [{user_id}]: Response delivered via WebSocket for '{subject}'")
            else:
                await self._pending.save(user_id, request_id, conversation_id, response_params)
                log.info(f"VirtualLoopback [{user_id}]: Client offline — result stored for '{subject}'")

        except Exception as e:
            log.error(f"VirtualLoopback [{user_id}]: Processing failed: {e}", exc_info=True)

            # Send error response (or store if client is offline)
            error_params = {
                "request_id": request_id,
                "subject": subject,
                "conversation_id": conversation_id,
                "original_message_id": original_message_id,
                "response_text": "An error occurred while processing your request. Please try again.",
                "response_html": "<p style='color:red'>Error: unable to process your request. Please try again.</p>",
            }
            ws = self.active_connections.get(user_id)
            if ws:
                try:
                    await ws.send_text(json.dumps({"jsonrpc": "2.0",
                                                   "method": "virtual_loopback/response",
                                                   "params": error_params}))
                except Exception:
                    log.error(f"VirtualLoopback [{user_id}]: Failed to send error response")
            else:
                await self._pending.save(user_id, request_id, conversation_id, error_params)

    def _process_attachment(self, att_data: dict) -> str:
        """Decode base64 attachment and extract text content."""
        file_name = att_data.get("file_name", "unknown")
        content_b64 = att_data.get("content_base64", "")

        if not content_b64:
            return f"[Empty attachment: {file_name}]"

        # Guard: reject oversized base64 before allocating decoded bytes in memory
        from src.config import VSTO_MAX_ATTACHMENT_MB
        max_encoded_len = int(VSTO_MAX_ATTACHMENT_MB * 1024 * 1024 * 1.34)
        if len(content_b64) > max_encoded_len:
            log.warning(f"VirtualLoopback: attachment '{file_name}' exceeds size guard ({VSTO_MAX_ATTACHMENT_MB} MB), skipping decode")
            return f"[Attachment too large to process server-side: {file_name}]"

        file_bytes = base64.b64decode(content_b64)
        ext = Path(file_name).suffix.lower()

        # Save to temporary file for processing
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False, prefix="arcumai_lb_") as tmp:
            tmp_path = Path(tmp.name)  # assign before write so finally cleanup always works
            tmp.write(file_bytes)

        try:
            if ext == ".pdf":
                from src.readers import SmartPDFReader
                reader = SmartPDFReader()
                docs = reader.load_data(tmp_path)
                return "\n".join([d.text for d in docs])

            elif ext == ".msg":
                from src.readers import MyOutlookReader
                reader = MyOutlookReader()
                docs = reader.load_data(tmp_path)
                return "\n".join([d.text for d in docs])

            elif ext == ".eml":
                from src.readers import MyEmlReader
                reader = MyEmlReader()
                docs = reader.load_data(tmp_path)
                return "\n".join([d.text for d in docs])

            elif ext in (".txt", ".csv", ".md"):
                return file_bytes.decode("utf-8", errors="ignore")

            elif ext == ".docx":
                try:
                    import docx
                    doc = docx.Document(tmp_path)
                    return "\n".join([p.text for p in doc.paragraphs])
                except ImportError:
                    return f"[DOCX support requires python-docx: {file_name}]"

            elif ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(tmp_path, data_only=True)
                    text_parts = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        text_parts.append(f"[Sheet: {sheet_name}]")
                        for row in ws.iter_rows(values_only=True):
                            row_text = " | ".join([str(c) if c is not None else "" for c in row])
                            if row_text.strip():
                                text_parts.append(row_text)
                    return "\n".join(text_parts)
                except ImportError:
                    return f"[Excel support requires openpyxl: {file_name}]"

            else:
                return f"[Unsupported file type: {ext}]"

        finally:
            try:
                tmp_path.unlink()
            except Exception:
                pass

    async def _route_to_ai_engine(
        self,
        user_id: str,
        subject: str,
        body: str,
        attachment_context: str = "",
        use_rag: bool = True
    ) -> str:
        """Route the loopback request to the appropriate AI engine.

        For RAG (no attachments): optimize email body as a search query, search knowledge base.
        For FILE_READER (with attachments): use body as the question/instruction,
            attachment_context as the document content to analyse.
        """
        from src.auth import load_users
        from src.engine import UserSession, optimize_prompt_for_rag

        # Find the ArcumAI username from the outlook_id
        users = load_users()
        username = None
        role = "DEFAULT"
        for uname, udata in users.items():
            if udata.get("outlook_id") == user_id:
                username = uname
                role = udata.get("role", "DEFAULT")
                break

        if not username:
            return f"Error: No ArcumAI user associated with outlook_id '{user_id}'."

        # Create a temporary session for this request
        session = UserSession(username=username, role=role)

        if use_rag:
            # No attachments: optimise the email body as a RAG search query
            log.info(f"VirtualLoopback [{user_id}]: Incoming email | Subject='{subject}' | Body ({len(body)} chars):\n{body}")

            try:
                optimized_query = await optimize_prompt_for_rag(subject, body)
                log.info(f"VirtualLoopback [{user_id}]: Optimized prompt:\n{optimized_query}")
            except Exception as e:
                log.warning(f"VirtualLoopback [{user_id}]: Prompt optimization failed ({e}), using raw email as fallback")
                optimized_query = f"Email Subject: {subject}\n\n{body}"

            query = f"@rag {optimized_query}"
            mode_override = None
            log.info(f"VirtualLoopback [{user_id}]: Routing to RAG engine (no attachments)")
        else:
            # Has attachments: body = question, attachment_context = document text
            # The body goes through optimize_prompt_for_rag to strip email noise and extract the
            # core instruction (NER masking also applied if configured).
            log.info(f"VirtualLoopback [{user_id}]: Incoming email with attachments | Subject='{subject}' | "
                     f"Body ({len(body)} chars) | Attachment context ({len(attachment_context)} chars)")
            try:
                optimized_instruction = await optimize_prompt_for_rag(subject, body)
                log.info(f"VirtualLoopback [{user_id}]: Optimized instruction:\n{optimized_instruction}")
            except Exception as e:
                log.warning(f"VirtualLoopback [{user_id}]: Prompt optimization failed ({e}), using raw body as fallback")
                optimized_instruction = f"Subject: {subject}\n\n{body}"

            query = optimized_instruction          # What the user wants done with the document(s)
            session.uploaded_context = attachment_context  # Document text only (no body duplication)
            mode_override = "FILE_READER"
            log.info(f"VirtualLoopback [{user_id}]: Routing to FILE_READER engine "
                     f"(instruction: {len(query)} chars | context: {len(attachment_context)} chars)")

        try:
            response_obj, response_text, used_mode = await session.run_chat_action(query, mode_override=mode_override)
            log.info(f"VirtualLoopback [{user_id}]: AI responded via {used_mode} mode")
            return response_text
        except Exception as e:
            log.error(f"VirtualLoopback [{user_id}]: AI engine error: {e}", exc_info=True)
            return f"Error during AI processing: {str(e)}"

    def _build_cc_disclaimer(self, cc_recipients: list) -> str:
        """Build a disclaimer for emails where ArcumAI was in CC with real recipients."""
        names = ", ".join(cc_recipients)
        return (
            "------------------------------------------------------------\n"
            "NOTE: This response is for you only.\n"
            f"The CC recipients of the original email ({names}) "
            "did NOT receive this analysis.\n"
            "If you find it useful, you can forward this email to them.\n"
            "------------------------------------------------------------"
        )

    def _markdown_to_html(self, markdown_text: str) -> str:
        """Convert markdown text to HTML for email display."""
        try:
            import markdown
            return markdown.markdown(
                markdown_text,
                extensions=['tables', 'fenced_code', 'nl2br']
            )
        except ImportError:
            # Fallback: basic conversion
            import html as html_module
            escaped = html_module.escape(markdown_text)
            return escaped.replace('\n', '<br>')
